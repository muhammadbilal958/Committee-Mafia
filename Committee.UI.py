from flask import Flask, render_template, request, redirect, url_for, flash, send_file
import json, os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime
from io import BytesIO

app = Flask(__name__)
app.secret_key = "fast_fintech_bilal_ultimate_v8"
json_file = "committee_data.json"

def load_data():
    if os.path.exists(json_file):
        with open(json_file, "r") as f:
            try:
                data = json.load(f)
                if isinstance(data, dict) and "members" in data:
                    members = {int(k): v for k, v in data["members"].items()}
                    pool = data.get("penalty_pool", 0)
                    return {"members": members, "penalty_pool": pool}
                else:
                    members = {int(k): v for k, v in data.items()}
                    return {"members": members, "penalty_pool": 0}
            except: return {"members": {}, "penalty_pool": 0}
    return {"members": {}, "penalty_pool": 0}

def save_data(data_to_save):
    with open(json_file, "w") as f:
        json.dump(data_to_save, f, indent=4)

def get_current_collection(members):
    all_payments = sum(m['balance'] for m in members.values())
    done_members = sum(1 for m in members.values() if m.get('committee_status') == 'Done')
    target_per_round = len(members) * 5000 if members else 5000
    return all_payments - (done_members * target_per_round)

@app.route('/')
def index():
    data = load_data()
    members = data["members"]
    pool = data["penalty_pool"]
    current_collection = get_current_collection(members)
    sorted_members = dict(sorted(members.items(), key=lambda x: x[1]['score'], reverse=True))
    bonus = pool / len(members) if members else 0
    return render_template('index.html', members=sorted_members, total=current_collection, penalty_pool=pool, bonus=round(bonus, 2))

@app.route('/register', methods=['POST'])
def register():
    data = load_data()
    m_id = int(request.form['id'])
    if m_id in data["members"]:
        flash("ID pehle se mojood hai!", "danger")
    else:
        data["members"][m_id] = {
            'Member_Name': request.form['name'], 'Member_id': m_id, 
            'Member_Gurantor': request.form['g_name'], 'gurantor_id': int(request.form['g_id']), 
            'score': 100, 'balance': 0, 'history': {}, 'committee_status': 'Not Done',
            'is_paid': False # <-- FIX: Backend crash nahi karega
        }
        save_data(data)
        flash("New Member Registered!", "success")
    return redirect(url_for('index'))

@app.route('/pay', methods=['POST'])
def pay():
    data = load_data()
    members = data["members"]
    m_id = int(request.form['m_id'])
    mon_num = int(request.form['month'])
    date = int(request.form['date'])
    months = ["Jan", "Feb", "March", "April", "May", "June", "July", "Aug", "Sept", "Oct", "Nov", "Dec"]
    
    if m_id in members:
        m_name = months[mon_num-1]
        if m_name in members[m_id]['history']:
            flash(f"Error: {m_name} already paid!", "danger")
            return redirect(url_for('index'))
        
        p_type = "Late" if date > 5 else "ontime"
        if p_type == "Late":
            members[m_id]['score'] -= 10
            gid = members[m_id]['gurantor_id']
            if gid in members: members[gid]['score'] -= 3
        else: 
            members[m_id]['score'] += 5
            
        members[m_id]['balance'] += 5000
        members[m_id]['history'][m_name] = {"payment_type": p_type}
        members[m_id]['is_paid'] = True # <-- FIX: Backend summary ko 'is_paid' mil jayega
        
        # --- AUTO PAYOUT LOGIC ---
        curr_coll = get_current_collection(members)
        target = len(members) * 5000
        if curr_coll >= target:
            winner_id = None
            max_s = -9999
            for pid, pdata in members.items():
                if pdata.get('committee_status') != "Done" and pdata['score'] > max_s:
                    max_s = pdata['score']
                    winner_id = pid
            if winner_id:
                members[winner_id]['committee_status'] = "Done"
                members[winner_id]['payout_date'] = datetime.now().strftime("%d-%b-%Y")
                flash(f"🎊 WINNER: {members[winner_id]['Member_Name']} won!", "info")

        save_data(data)
        flash(f"Payment for {m_name} recorded!", "success")
    return redirect(url_for('index'))

@app.route('/delete/<int:m_id>/<month>')
def delete_payment(m_id, month):
    data = load_data()
    if m_id in data["members"] and month in data["members"][m_id]['history']:
        data["members"][m_id]['balance'] -= 5000
        del data["members"][m_id]['history'][month]
        # Agar koi history na bache to is_paid False kar do
        if not data["members"][m_id]['history']:
            data["members"][m_id]['is_paid'] = False 
        save_data(data); flash("Payment deleted", "info")
    return redirect(url_for('index'))

@app.route('/delete_member/<int:m_id>')
def delete_member(m_id):
    data = load_data()
    if m_id in data["members"]:
        bal = data["members"][m_id]['balance']
        if bal > 0:
            data["penalty_pool"] += (bal * 0.5)
        del data["members"][m_id]
        save_data(data); flash("Member deleted with penalty", "warning")
    return redirect(url_for('index'))

@app.route('/reset', methods=['POST'])
def reset():
    if request.form.get('confirm', '').upper() == "YES":
        save_data({"members": {}, "penalty_pool": 0})
        flash("SYSTEM RESET!", "success")
    return redirect(url_for('index'))

@app.route('/export')
def export():
    data = load_data()
    wb = Workbook(); ws = wb.active
    ws.title = "Committee_Report"
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    headers = ["Rank", "Name", "Score", "Balance", "Status", "Payout Date"]
    for col, val in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=val)
        cell.fill = header_fill; cell.font = Font(color="FFFFFF", bold=True)
    
    sorted_m = sorted(data["members"].values(), key=lambda x: x['score'], reverse=True)
    for i, d in enumerate(sorted_m, 2):
        ws.cell(row=i, column=1, value=i-1)
        ws.cell(row=i, column=2, value=d['Member_Name'])
        ws.cell(row=i, column=3, value=d['score'])
        ws.cell(row=i, column=4, value=d['balance'])
        ws.cell(row=i, column=5, value=d['committee_status'])
        ws.cell(row=i, column=6, value=d.get('payout_date', '-'))
    
    output = BytesIO(); wb.save(output); output.seek(0)
    return send_file(output, as_attachment=True, download_name="Committee_Report.xlsx")

if __name__ == '__main__':
    app.run(debug=True)