import json
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime

# File names
json_file = "committee_data.json"
excel_file = "Committee_Master_Report.xlsx"

def generate_pro_report():
    if not os.path.exists(json_file):
        print(f"Error: '{json_file}' nahi mili!")
        return

    # 1. JSON Data Load (With Format Fix)
    with open(json_file, "r") as file:
        data = json.load(file)
        
        # FIX: Check format to avoid ValueError
        if isinstance(data, dict) and "members" in data:
            member_list = {int(k): v for k, v in data["members"].items()}
            current_penalty_pool = data.get("penalty_pool", 0)
        else:
            member_list = {int(k): v for k, v in data.items()}
            current_penalty_pool = 0

    # --- UPDATED PAYOUT STATUS LOGIC ---
    target_payout = len(member_list) * 5000
    # Sirf unka balance jo abhi list mein hain
    total_received = sum(m.get('balance', 0) for m in member_list.values())
    done_count = sum(1 for m in member_list.values() if m.get('committee_status') == 'Done')
    
    # Collection mein se payout nikal kar jo bacha
    current_pool = total_received - (done_count * target_payout)

    # Winner selection logic
    if current_pool >= target_payout:
        winner_id = None
        max_score = -1001
        
        for mid, mdata in member_list.items():
            if mdata.get('committee_status') != "Done":
                if mdata['score'] > max_score:
                    max_score = mdata['score']
                    winner_id = mid
        
        if winner_id:
            member_list[winner_id]['committee_status'] = "Done"
            member_list[winner_id]['payout_date'] = datetime.now().strftime("%d-%b-%Y")
            print(f"🎊 AUTO-PAYOUT: {member_list[winner_id]['Member_Name']} is the winner!")
            
            # Save back with new format
            with open(json_file, "w") as file:
                json.dump({"members": member_list, "penalty_pool": current_penalty_pool}, file, indent=4)
            
            current_pool -= target_payout

    # 2. Excel Setup
    try:
        if os.path.exists(excel_file):
            wb = load_workbook(excel_file)
        else:
            wb = Workbook()
            if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])
            wb.create_sheet("Current_Committee")
    except PermissionError:
        print("\n[ERROR] Excel khuli hui hai! Band kar ke run karein.")
        return

    ws = wb["Current_Committee"]
    ws.delete_rows(1, ws.max_row + 10) # Purana data saaf karne ke liye

    # 3. Styles
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    gold_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    pool_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    penalty_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    
    months_map = ["Jan", "Feb", "March", "April", "May", "June", "July", "Aug", "Sept", "Oct", "Nov", "Dec"]
    headers = ["Rank", "Member Name", "Score", "Total Paid", "Status", "Payout Date"] + months_map + ["Pending", "Risk Alert"]
    
    for col, val in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=val)
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    # 4. Data Insertion
    sorted_members = sorted(member_list.values(), key=lambda x: x['score'], reverse=True)
    current_month_idx = datetime.now().month
    last_row_idx = 1

    for index, d in enumerate(sorted_members, start=2):
        paid_history = d.get("history", {})
        pending_count = 0
        member_balance = d.get('balance', 0)
        status_done = d.get("committee_status", "Not Done")
        payout_dt = d.get("payout_date", "-")

        row_data = [index-1, d['Member_Name'], d['score'], member_balance, status_done, payout_dt]
        
        for i in range(12):
            m_name = months_map[i]
            if m_name in paid_history:
                row_data.append("PAID")
            elif i < current_month_idx:
                row_data.append("UNPAID")
                pending_count += 1
            else:
                row_data.append("-")
        
        alert = "Safe"
        if 1 <= pending_count <= 2: alert = f"Late ({pending_count})"
        elif 3 <= pending_count <= 4: alert = "CRITICAL: Guarantor Alert!"
        elif pending_count >= 5: alert = "TERMINATED"
        row_data.extend([pending_count, alert])

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=index, column=col, value=val)
            cell.alignment = Alignment(horizontal="center")
            if val == "PAID": cell.fill = green_fill
            elif val == "UNPAID": cell.fill = red_fill
            if status_done == "Done" and col <= 6: cell.fill = gold_fill
        
        last_row_idx = index

    # --- SUMMARY SECTION (Updated with Penalty Pool) ---
    summary_start = last_row_idx + 3
    
    # Pool Collection Row
    ws.cell(row=summary_start, column=2, value="CURRENT COMMITTEE COLLECTION:").font = Font(bold=True)
    p_cell = ws.cell(row=summary_start, column=4, value=f"Rs. {current_pool}")
    p_cell.fill = pool_fill
    p_cell.font = Font(bold=True)

    # Penalty Pool Row
    ws.cell(row=summary_start + 1, column=2, value="PENALTY POOL (FOR BONUS):").font = Font(bold=True)
    pen_cell = ws.cell(row=summary_start + 1, column=4, value=f"Rs. {current_penalty_pool}")
    pen_cell.fill = penalty_fill
    pen_cell.font = Font(bold=True)
    
    # Extra Share Info
    if len(member_list) > 0:
        share = current_penalty_pool / len(member_list)
        ws.cell(row=summary_start + 2, column=2, value="ESTIMATED BONUS PER MEMBER:").font = Font(italic=True)
        ws.cell(row=summary_start + 2, column=4, value=f"Rs. {round(share, 2)}").font = Font(italic=True)

    # Final Save
    wb.save(excel_file)
    print(f"✅ Excel Updated! Pool: Rs. {current_pool} | Penalty: Rs. {current_penalty_pool}")

if __name__ == "__main__":
    generate_pro_report()